import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelWriter:
    def __init__(self, filename, dest_path):
        self.filename = filename
        self.dest_path = dest_path
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def add_sheet(self, name):
        self.ws = self.wb.create_sheet(title=name)

    def write_data(self, data):
        try:
            for r in dataframe_to_rows(data, index=False, header=True):
                self.ws.append(r)
        except Exception as e:
            print(f"Error writing data to Excel file: {e}")
            return False
        return True

    def save(self):
        try:
            self.wb.save(f"{self.dest_path}/{self.filename}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False
        return True

    def append_all_sheets(self, new_sheet_name):
        try:
            wb = openpyxl.load_workbook(f"{self.dest_path}/{self.filename}")
            new_sheet = wb.create_sheet(title=new_sheet_name)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # transpose the sheet and add to the new sheet
                for column in sheet.iter_cols():
                    new_sheet.append([cell.value for cell in column])
            wb.save(f"{self.dest_path}/{self.filename}")
        except Exception as e:
            print(f"Error appending sheets to Excel file: {e}")
            return False
        return True
    # def append_all_sheets(self, new_sheet_name):
    #     try:
    #         wb = openpyxl.load_workbook(f"{self.dest_path}/{self.filename}")
    #         new_sheet = wb.create_sheet(title=new_sheet_name)
    #         for sheet_name in wb.sheetnames:
    #             sheet = wb[sheet_name]
    #             for row in sheet.iter_rows():
    #                 new_sheet.append([cell.value for cell in row])
    #         wb.save(f"{self.dest_path}/{self.filename}")
    #     except Exception as e:
    #         print(f"Error appending sheets to Excel file: {e}")
    #         return False
    #     return True