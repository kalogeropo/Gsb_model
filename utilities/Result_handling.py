import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from openpyxl import Workbook
from typing import Optional
from pandas import DataFrame



def res_to_excel(result_model, namefile='example.xlsx', dest_path="collections/test/Results", sheetname="test"):
    df = result_model.results_to_df()
    write(xl_namefile=namefile, dest_path=dest_path, sheetname=sheetname, data=df)

def expir_start(path,path_to_write,col_path):

    from Preprocess.Collection import Collection
    testcol = Collection(path, name="test")
    testcol.create_collection()
    testcol.save_inverted_index(path_to_write)
    r, q = testcol.load_collection(col_path)
    return testcol,q,r

def write(xl_namefile='example.xlsx', dest_path="collections/test/Results", sheetname="test", data=None):
    writer = ExcelWriter(xl_namefile, dest_path)
    writer.add_sheet(sheetname)
    writer.write_data(data)
    writer.save()
    # writer.append_all_sheets("all_tests")


class ExcelWriter:
    """
    A utility class for reading, writing, and managing Excel files using openpyxl.
    """
    def __init__(self, filename: str, dest_path: str ):
        self.filename = filename
        self.dest_path = dest_path
        file_path = os.path.join(self.dest_path, self.filename)

        if os.path.exists(file_path):
            try:
                self.wb = load_workbook(file_path)
            except Exception as e:
                print(f"Error loading workbook, creating new one: {e}")
                self.wb = Workbook()
        else:
            self.wb = Workbook()

        self.ws = self.wb.active

    def add_sheet(self, name):
        self.ws = self.wb.create_sheet(title=name)

    def write_data(self, data: Optional[DataFrame]) -> bool:  
        """
        Writes a Pandas DataFrame to an Excel worksheet.
        Args:
            data (pd.DataFrame): The data to write to Excel. Must not be None.

        Returns:
            bool: True if writing succeeded, False otherwise.
        """
        if data is None:
            print("No data provided to write.")
            return False

        if self.ws is None:
            print("Invalid or uninitialized worksheet.")
            return False
        try:
            for row in dataframe_to_rows(data, index=False, header=True):
                self.ws.append(row)
            return True
        except Exception as e:
            print(f"Error writing data to Excel file: {e}")
        return False

    def save(self) -> bool:
        """
        Saves the workbook to the current file path.

        Returns:
            bool: True if saved successfully, False otherwise.
        """
        try:
            self.wb.save(os.path.join(self.dest_path, self.filename))
            return True
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False

    def append_all_sheets(self, new_sheet_name):
        try:
            wb = openpyxl.load_workbook(f"{self.dest_path}/{self.filename}")
            new_sheet = wb.create_sheet(title=new_sheet_name)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # transpose the sheet and add to the new sheet
                for column in sheet.iter_cols():
                    new_sheet.append([cell.value for cell in column])
            wb.save(f"{self.dest_path}/{self.filename}")
        except Exception as e:
            print(f"Error appending sheets to Excel file: {e}")
            return False
        return True
    # def append_all_sheets(self, new_sheet_name):
    #     try:
    #         wb = openpyxl.load_workbook(f"{self.dest_path}/{self.filename}")
    #         new_sheet = wb.create_sheet(title=new_sheet_name)
    #         for sheet_name in wb.sheetnames:
    #             sheet = wb[sheet_name]
    #             for row in sheet.iter_rows():
    #                 new_sheet.append([cell.value for cell in row])
    #         wb.save(f"{self.dest_path}/{self.filename}")
    #     except Exception as e:
    #         print(f"Error appending sheets to Excel file: {e}")
    #         return False
    #     return True
